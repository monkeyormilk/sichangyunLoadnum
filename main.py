import openpyxl
import os
import shutil
import datetime
import sqldb


#将每日的日报复制到指定目录
def copyfile(year, month):
    #每日日报存放路径
    sourcePath = r'D:\onedrive2\OneDrive\各省数据\NJSCY各省每日日报'
    #复制到目标路径
    targetPath = r'D:\onedrive2\OneDrive\各省数据\重庆\各省日报汇总'
    # 复制各省日报
    for y in range(2):
        numcq = numhb = numsx = numlj = 0
        pyear = 2021 + y
        for m in range(month):
            cqname = "重庆" + str(month - m) + "月全月日报-" + str(pyear) + "年" + str(month - m) + "月.xlsx"
            cqfilepath = sourcePath + '/' + cqname
            if os.path.exists(cqfilepath) and numcq == 0:
                print('正在复制：重庆日报')
                shutil.copy(cqfilepath, targetPath + '/重庆/' + cqname)
                numcq += 1
                break
        for m in range(month):
            sxname = "山西联通－运营部日报" + str(month - m) + "月.xlsx"
            sxfilepath = sourcePath + '/' + sxname
            if os.path.exists(sxfilepath) and numsx == 0:
                print('正在复制：山西日报')
                shutil.copy(sxfilepath, targetPath + '/山西/' + sxname)
                numsx += 1
                break
        for m in range(month):
            hbname = "河北联通－运营部日报" + str(month - m) + "月.xlsx"
            hbfilepath = sourcePath + '/' + hbname
            if os.path.exists(hbfilepath) and numhb == 0:
                print('正在复制：河北日报')
                shutil.copy(hbfilepath, targetPath + '/河北/' + hbname)
                numhb += 1
                break
        for m in range(month):
            hljname = "黑龙江联通日报" + str(month - m) + "月.xlsx"
            ljfilepath = sourcePath + '/' + hljname
            if os.path.exists(ljfilepath) and numlj == 0:
                print('正在复制：龙江日报')
                shutil.copy(ljfilepath, targetPath + '/黑龙江/' + hljname)
                numlj += 1
                break
        if numcq == 1 and numhb ==1 and numsx == 1 and numlj == 1:
            break


#读取日报数据
def readnum(year, month):
    # day = datetime.datetime.now().day

    # if month == 1 and day == 1:
    #     year = year - 1
    # else:
    #     year = year
    # if day == 1:
    #     month = month - 1
    readcq(year, month)
    readhb(year, month)
    readsx(year, month)
    readhlj(year, month)


#读取重庆日报数据，并保存至数据库
def readcq(year, month):
    # 存放数据
    # 存放每天订购量
    day_sale = []
    # 存放每天佣金
    day_money = []
    # 每月截至当天佣金合计
    total_money = []
    # 存放月包销量
    yuebao_sale = []
    # 存放每天多日包订购量
    duoribao_sale = []
    # 存放5G升级包订购量
    shengjibao_sale = []
    # 存放日包订购量
    ribao_sale = []
    # 存放语音包订购量
    yuyinbao_sale = []
    # 存放发起量
    day_faqi = []
    # 存放订购量
    day_dinggou = []
    # 存放月包佣金
    yuebao_money = []
    # 存放多日包佣金
    duoribao_money = []
    # 存放日包佣金
    ribao_money = []
    # 存放5G升级包佣金
    shengjibao_money = []
    # 存放语音包佣金
    yuyinbao_money = []

    # 读取重庆数据
    CQpath = r'D:\onedrive2\OneDrive\各省数据\重庆\各省日报汇总\重庆'
    #遍历每个日报文件
    for m in range(month):
        year = str(year)
        m = str(m + 1)
        CQfileName = CQpath + '/' + "重庆" + m + "月全月日报-" + year + "年" + m + "月.xlsx"
        if int(m) == month:
            if not os.path.exists(CQfileName):
                print('不存在')
                break
        print('正在读取：' + CQfileName)
        # 读取日报,data_only=True读取单元格的值 而不是公式
        excel = openpyxl.load_workbook(CQfileName, data_only=True)
        # 获取日报sheet名
        # names = excel.sheetnames
        # print(names)
        sheet = excel['销量与收入']
        for i in range(31):
            # 存入每天订购量
            value = sheet.cell(row=i + 4, column=2).value
            day_sale.append(value)
            # 存入每天佣金
            money = round(sheet.cell(row=i + 4, column=4).value, 1)
            day_money.append(money)
            # 存入截至当日总佣金
            day_totalmoney = round(sheet.cell(row=i + 4, column=6).value, 1)
            total_money.append(day_totalmoney)
            # 存入月包订购量
            yuebaos = sheet.cell(row=i + 4, column=11).value
            yuebao_sale.append(yuebaos)
            # 存入每天多日包订购量
            drb_sale = sheet.cell(row=i + 4, column=12).value
            duoribao_sale.append(drb_sale)
            # 存放5G升级包订购量
            shengjibao = sheet.cell(row=i + 4, column=13).value
            shengjibao_sale.append(shengjibao)
            # 存放日包订购量
            ribao = sheet.cell(row=i + 4, column=14).value
            ribao_sale.append(ribao)
            # 存放语音包订购量
            if int(m) > 2:
                yuyinbao = sheet.cell(row=i + 4, column=15).value
                yuyinbao_sale.append(yuyinbao)
            else:
                yuyinbao_sale.append(0)

        sheet = excel['发起订购']
        for i in range(31):
            if int(m) == 1:
                faqi = sheet.cell(row=61, column=i + 4).value
            elif int(m) == 2:
                faqi = sheet.cell(row=68, column=i + 4).value
            elif int(m) == 3:
                faqi = sheet.cell(row=76, column=i + 4).value
            elif int(m) > 7:
                faqi = sheet.cell(row=47, column=i + 4).value
            else:
                faqi = sheet.cell(row=63, column=i + 4).value
            day_faqi.append(faqi)

        sheet = excel['订购成功']
        for i in range(31):
            if int(m) == 1:
                dinggou = sheet.cell(row=61, column=i + 10).value
                yuebaom = sheet.cell(row=76, column=i + 10).value
                duoribaom = sheet.cell(row=77, column=i + 10).value
                shengjibaom = sheet.cell(row=78, column=i + 10).value
                ribaom = sheet.cell(row=79, column=i + 10).value
                yuyinbaom = 0
            elif int(m) == 2:
                dinggou = sheet.cell(row=68, column=i + 10).value
                yuebaom = sheet.cell(row=84, column=i + 10).value
                duoribaom = sheet.cell(row=85, column=i + 10).value
                shengjibaom = sheet.cell(row=86, column=i + 10).value
                ribaom = sheet.cell(row=87, column=i + 10).value
                yuyinbaom = sheet.cell(row=88, column=i + 10).value
            elif int(m) == 3:
                dinggou = sheet.cell(row=76, column=i + 10).value
                yuebaom = sheet.cell(row=92, column=i + 10).value
                duoribaom = sheet.cell(row=93, column=i + 10).value
                shengjibaom = sheet.cell(row=94, column=i + 10).value
                ribaom = sheet.cell(row=95, column=i + 10).value
                yuyinbaom = sheet.cell(row=96, column=i + 10).value
            elif int(m) > 7:
                dinggou = sheet.cell(row=47, column=i + 10).value
                yuebaom = sheet.cell(row=63, column=i + 10).value
                duoribaom = sheet.cell(row=64, column=i + 10).value
                shengjibaom = sheet.cell(row=65, column=i + 10).value
                ribaom = sheet.cell(row=66, column=i + 10).value
                yuyinbaom = sheet.cell(row=67, column=i + 10).value
            else:
                dinggou = sheet.cell(row=63, column=i + 10).value
                yuebaom = sheet.cell(row=79, column=i + 10).value
                duoribaom = sheet.cell(row=80, column=i + 10).value
                shengjibaom = sheet.cell(row=81, column=i + 10).value
                ribaom = sheet.cell(row=82, column=i + 10).value
                yuyinbaom = sheet.cell(row=83, column=i + 10).value
            yuebaom = round(yuebaom, 1)
            duoribaom = round(duoribaom, 1)
            shengjibaom = round(shengjibaom, 1)
            ribaom = round(ribaom, 1)
            yuyinbaom = round(yuyinbaom, 1)
            day_dinggou.append(dinggou)
            yuebao_money.append(yuebaom)
            duoribao_money.append(duoribaom)
            shengjibao_money.append(shengjibaom)
            ribao_money.append(ribaom)
            yuyinbao_money.append(yuyinbaom)
        excel.close()
    # 插入至数据库
    if not os.path.exists(CQfileName):
        month -= 1
    sqldb.insertcq(year, month, day_sale, day_money, total_money, yuebao_sale, duoribao_sale, shengjibao_sale, ribao_sale,
                    yuyinbao_sale, day_faqi, day_dinggou, yuebao_money, duoribao_money,shengjibao_money,ribao_money,yuyinbao_money)


#读取河北日报数据
def readhb(year, month):
    # 存放数据
    # 存放每天订购量
    day_sale = []
    # 存放每天佣金
    day_money = []
    # 每月截至当天佣金合计
    total_money = []
    # 特惠月包每天订购量
    tehuiyuebao_sale = []
    # 存放每天多日包订购量
    duoribao_sale = []
    # 存放5G升级包订购量
    shengjibao_sale = []
    # 存放语音包订购量
    yuyinbao_sale = []
    # 存放权益包订购量
    quanyibao_sale = []
    # 存放加速包订购量
    jiasubao_sale = []
    # 月包订购量
    yuebao_sale = []
    # 闲时包
    xianshibao_sale = []
    # 存放发起量
    day_faqi = []
    # 存放订购量
    day_dinggou = []
    # 二次确认量
    day_erciqueren = []
    # 存放月包佣金
    yuebao_money = []
    # 存放特惠月包佣金
    tehui_money = []
    # 存放日包佣金
    ribao_money = []
    # 闲时包佣金
    xianshibao_money = []
    # 权益包佣金
    quanyibao_money = []
    # 加速包佣金
    jiasubao_money = []
    # 升级包佣金
    shengjibao_money = []
    # 语音包佣金
    yuyinbao_money = []

    # 读取河北数据
    HBpath = r'D:\onedrive2\OneDrive\各省数据\重庆\各省日报汇总\河北'
    #遍历每个日报文件
    for m in range(month):
        year = str(year)
        m = str(m + 1)
        HBfileName = HBpath + '/' + "河北联通－运营部日报" + m + "月.xlsx"
        if int(m) == month:
            if not os.path.exists(HBfileName):
                print('不存在')
                break
        print('正在读取：' + HBfileName)
        # 读取日报,data_only=True读取单元格的值 而不是公式
        excel = openpyxl.load_workbook(HBfileName, data_only=True)
        # 获取日报sheet名
        # names = excel.sheetnames
        # print(names)
        sheet = excel['销量与收入']

        for i in range(31):
            # 存入每天订购量
            value = sheet.cell(row=i + 4, column=2).value
            day_sale.append(value)
            # 存入每天佣金
            money = round(sheet.cell(row=i + 4, column=4).value, 1)
            day_money.append(money)
            # 存入截至当日总佣金
            day_totalmoney = round(sheet.cell(row=i + 4, column=6).value, 1)
            total_money.append(day_totalmoney)
            #存入特惠月包
            tehuiyuebao = sheet.cell(row=i+4, column=11).value
            tehuiyuebao_sale.append(tehuiyuebao)
            # 存入每天多日包订购量
            drb_sale = sheet.cell(row=i + 4, column=12).value
            duoribao_sale.append(drb_sale)
            # 存放语音包订购量
            yuyinbao = sheet.cell(row=i + 4, column=13).value
            yuyinbao_sale.append(yuyinbao)
            # 存放权益包订购量
            quanyibao = sheet.cell(row=i + 4, column=14).value
            quanyibao_sale.append(quanyibao)
            # 存放5G升级包订购量
            shengjibao = sheet.cell(row=i + 4, column=15).value
            shengjibao_sale.append(shengjibao)
            # 存放加速包订购量
            jiasubao = sheet.cell(row=i + 4, column=16).value
            jiasubao_sale.append(jiasubao)
            # 存放月包订购量
            yuebao = sheet.cell(row=i + 4, column=17).value
            yuebao_sale.append(yuebao)
            # 存放闲时包
            xianshibao = sheet.cell(row=i + 4, column=18).value
            xianshibao_sale.append(xianshibao)

        sheet = excel['发起订购']
        for i in range(31):
            if int(m) > 6:
                faqi = sheet.cell(row=37, column=i + 5).value
            else:
                faqi = sheet.cell(row=36, column=i + 5).value
            day_faqi.append(faqi)

        sheet = excel['订购成功']
        for i in range(31):
            if int(m) > 6:
                dinggou = sheet.cell(row=37, column=i + 10).value
            else:
                dinggou = sheet.cell(row=36, column=i + 10).value

            day_dinggou.append(dinggou)

        sheet = excel['二次确认']
        for i in range(31):
            if int(m) > 6:
                erciqueren = sheet.cell(row=37, column=i + 5).value
            else:
                erciqueren = sheet.cell(row=36, column=i + 5).value
            day_erciqueren.append(erciqueren)

        sheet = excel['佣金结算表']
        # 存放各日产品佣金
        yuebaom = 0
        tehuim = 0
        ribaom = 0
        xianshibaom = 0
        quanyibaom = 0
        jiasubaom = 0
        shengjibaom = 0
        yuyinbaom = 0
        for i in range(31):
            if int(m) > 6:
                for r in range(35):
                    if r < 4:
                        yuebaom += sheet.cell(row=2 + r,column = 7 + i).value
                    elif r < 9:
                        tehuim += sheet.cell(row=2 + r,column = 7 + i).value
                    elif r < 15:
                        ribaom += sheet.cell(row=2 + r,column = 7 + i).value
                    elif r < 17:
                        xianshibaom += sheet.cell(row=2 + r,column = 7 + i).value
                    elif r < 25:
                        quanyibaom += sheet.cell(row=2 + r,column = 7 + i).value
                    elif r < 27:
                        jiasubaom += sheet.cell(row=2 + r,column = 7 + i).value
                    elif r < 32:
                        shengjibaom += sheet.cell(row=2 + r,column = 7 + i).value
                    else:
                        yuyinbaom += sheet.cell(row=2 + r,column = 7 + i).value
            else:
                for r in range(34):
                    if r < 4:
                        yuebaom += sheet.cell(row=2 + r,column = 7 + i).value
                    elif r < 9:
                        tehuim += sheet.cell(row=2 + r,column = 7 + i).value
                    elif r < 15:
                        ribaom += sheet.cell(row=2 + r,column = 7 + i).value
                    elif r < 17:
                        xianshibaom += sheet.cell(row=2 + r,column = 7 + i).value
                    elif r < 25:
                        quanyibaom += sheet.cell(row=2 + r,column = 7 + i).value
                    elif r < 26:
                        jiasubaom += sheet.cell(row=2 + r,column = 7 + i).value
                    elif r < 31:
                        shengjibaom += sheet.cell(row=2 + r,column = 7 + i).value
                    else:
                        yuyinbaom += sheet.cell(row=2 + r,column = 7 + i).value

            yuebaom = round(yuebaom, 1)
            tehuim = round(tehuim, 1)
            ribaom = round(ribaom, 1)
            xianshibaom = round(xianshibaom, 1)
            quanyibaom = round(quanyibaom, 1)
            jiasubaom = round(jiasubaom, 1)
            shengjibaom = round(shengjibaom, 1)
            yuyinbaom = round(yuyinbaom, 1)
            # 存放各产品累计佣金
            yuebao_money.append(yuebaom)
            tehui_money.append(tehuim)
            ribao_money.append(ribaom)
            xianshibao_money.append(xianshibaom)
            quanyibao_money.append(quanyibaom)
            jiasubao_money.append(jiasubaom)
            shengjibao_money.append(shengjibaom)
            yuyinbao_money.append(yuyinbaom)

        excel.close()
    if not os.path.exists(HBfileName):
        month -= 1
    sqldb.inserthb(year, month, day_sale, day_money, total_money, tehuiyuebao_sale, duoribao_sale, yuyinbao_sale, quanyibao_sale, shengjibao_sale, jiasubao_sale,
             yuebao_sale, xianshibao_sale, day_faqi,day_erciqueren, day_dinggou,yuebao_money,tehui_money,ribao_money,xianshibao_money,quanyibao_money,
             jiasubao_money,shengjibao_money,yuyinbao_money)


#读取山西日报数据
def readsx(year, month):
    # 存放数据
    # 存放每天订购量
    day_sale = []
    # 存放每天佣金
    day_money = []
    # 每月截至当天佣金合计
    total_money = []
    # 存放每天多日包订购量
    duoribao_sale = []
    # 存放5G升级包订购量
    shengjibao_sale = []
    # 存放语音包订购量
    yuyinbao_sale = []
    # 月包订购量
    yuebao_sale = []
    # 存放发起量
    day_faqi = []
    # 存放订购量
    day_dinggou = []
    # 二次确认量
    day_erciqueren = []
    # 日包多日包佣金
    duoribao_money = []
    # 月包佣金
    yuebao_money = []
    # 语音包佣金
    yuyinbao_money = []
    # 5G升级包佣金
    shengjibao_money = []
    # 2i类产品佣金
    shengdaibao_money = []

    # 读取河北数据
    sxpath = r'D:\onedrive2\OneDrive\各省数据\重庆\各省日报汇总\山西'
    #遍历每个日报文件
    for m in range(month):
        year = str(year)
        m = str(m + 1)
        sxfilename = sxpath + '/' + "山西联通－运营部日报" + m + "月.xlsx"
        if int(m) == month:
            if not os.path.exists(sxfilename):
                print('不存在')
                break
        print('正在读取：' + sxfilename)
        # 读取日报,data_only=True读取单元格的值 而不是公式
        excel = openpyxl.load_workbook(sxfilename, data_only=True)
        # 获取日报sheet名
        # names = excel.sheetnames
        # print(names)
        sheet = excel['销量与收入']

        for i in range(31):
            # 存入每天订购量
            value = sheet.cell(row=i + 4, column=2).value
            day_sale.append(value)
            # 存入每天佣金
            money = round(sheet.cell(row=i + 4, column=4).value, 1)
            day_money.append(money)
            # 存入截至当日总佣金
            day_totalmoney = round(sheet.cell(row=i + 4, column=6).value, 1)
            total_money.append(day_totalmoney)
            # 存入每天多日包订购量
            drb_sale = sheet.cell(row=i + 4, column=11).value
            duoribao_sale.append(drb_sale)
            # 存放语音包订购量
            yuyinbao = sheet.cell(row=i + 4, column=13).value
            yuyinbao_sale.append(yuyinbao)
            # 存放5G升级包订购量
            shengjibao = sheet.cell(row=i + 4, column=15).value
            shengjibao_sale.append(shengjibao)
            # 存放月包订购量
            yuebao = sheet.cell(row=i + 4, column=12).value
            yuebao_sale.append(yuebao)


        sheet = excel['发起订购']
        for i in range(31):
            faqi = sheet.cell(row=30, column=i + 5).value
            day_faqi.append(faqi)

        sheet = excel['订购成功']
        for i in range(31):
            dinggou = sheet.cell(row=30, column=i + 10).value
            day_dinggou.append(dinggou)

        sheet = excel['二次确认']
        for i in range(31):
            erciqueren = sheet.cell(row=30, column=i + 5).value
            day_erciqueren.append(erciqueren)

        sheet = excel['佣金结算表']

        duoribaom = 0
        yuebaom = 0
        yuyinbaom = 0
        shengjibaom = 0
        shengdaibaom = 0
        for i in range(31):
            for r in range(24):
                if r < 4:
                    duoribaom += sheet.cell(row=r + 2, column=i + 7).value
                elif r < 9:
                    yuebaom += sheet.cell(row=r + 2, column=i + 7).value
                elif r < 13:
                    yuyinbaom += sheet.cell(row=r + 2, column=i + 7).value
                elif r < 19:
                    shengjibaom += sheet.cell(row=r + 2, column=i + 7).value
                else:
                    shengdaibaom += sheet.cell(row=r + 2, column=i + 7).value
            duoribao_money.append(duoribaom)
            yuebao_money.append(yuebaom)
            yuyinbao_money.append(yuyinbaom)
            shengjibao_money.append(shengjibaom)
            shengdaibao_money.append(shengdaibaom)

        excel.close()
    if not os.path.exists(sxfilename):
        month -= 1
    sqldb.insertsx(year, month, day_sale, day_money, total_money, duoribao_sale, yuebao_sale, yuyinbao_sale,shengjibao_sale,day_faqi,day_erciqueren, day_dinggou,
                   duoribao_money, yuebao_money, yuyinbao_money, shengjibao_money, shengdaibao_money)


#读取黑龙江日报数据
def readhlj(year, month):
    # 存放数据
    # 存放每天订购量
    day_sale = []
    # 存放每天佣金
    day_money = []
    # 每月截至当天佣金合计
    total_money = []
    # 存放每天多日包订购量
    duoribao_sale = []
    # 存放圣代包订购量
    shengdaibao_sale = []
    # 超会订购量
    chaohui_sale = []
    # 存放语音包订购量
    yuyinbao_sale = []
    # 月包订购量
    yuebao_sale = []
    # 存放加速包订购量
    jiasubao_sale = []
    # 存放尊享包
    zunxiangbao_sale = []

    # 存放发起量
    day_faqi = []
    # 存放订购量
    day_dinggou = []
    # 二次确认量
    day_erciqueren = []
    # 存放多日包佣金
    duoribao_money = []
    # 冰激凌佣金
    bingjiling_money = []
    # 超会佣金
    chaohui_money = []
    # 语音包佣金
    yuyinbao_money = []
    # 月包佣金
    yuebao_money = []
    # 冰激凌加速包
    jiasubao_money = []
    # 尊享保佣金
    zunxiangbao_money = []
    # 王卡加速宝佣金

    # 读取河北数据
    hljpath = r'D:\onedrive2\OneDrive\各省数据\重庆\各省日报汇总\黑龙江'
    #遍历每个日报文件
    for m in range(month):
        year = str(year)
        m = str(m + 1)
        hljfilename = hljpath + '/' + "黑龙江联通日报" + m + "月.xlsx"
        if int(m) == month:
            if not os.path.exists(hljfilename):
                print('不存在')
                break
        print('正在读取：' + hljfilename)
        # 读取日报,data_only=True读取单元格的值 而不是公式
        excel = openpyxl.load_workbook(hljfilename, data_only=True)
        # 获取日报sheet名
        # names = excel.sheetnames
        # print(names)
        sheet = excel['销量与收入']

        for i in range(31):
            # 存入每天订购量
            value = sheet.cell(row=i + 4, column=2).value
            day_sale.append(value)
            # 存入每天佣金
            money = round(sheet.cell(row=i + 4, column=4).value, 1)
            day_money.append(money)
            # 存入截至当日总佣金
            day_totalmoney = round(sheet.cell(row=i + 4, column=6).value, 1)
            total_money.append(day_totalmoney)
            # 存入多日包
            duoribao = sheet.cell(row=i + 4, column=11).value
            duoribao_sale.append(duoribao)
            # 存入圣代包
            shengdaibao = sheet.cell(row=i + 4, column=12).value
            shengdaibao_sale.append(shengdaibao)
            # 存入超会
            chaohui = sheet.cell(row=i + 4, column=13).value
            chaohui_sale.append(chaohui)
            # 存入语音包
            yuyinbao = sheet.cell(row=i + 4, column=14).value
            yuyinbao_sale.append(yuyinbao)
            # 存入月包
            yuebao = sheet.cell(row=i + 4, column=15).value
            yuebao_sale.append(yuebao)
            # 存入加速包
            jiasubao = sheet.cell(row=i + 4, column=16).value
            jiasubao_sale.append(jiasubao)
            # 存入尊享包
            zunxiangbao = sheet.cell(row=i + 4, column=17).value
            zunxiangbao_sale.append(zunxiangbao)

        sheet = excel['发起订购']
        for i in range(31):
            faqi = sheet.cell(row=18, column=i + 4).value
            day_faqi.append(faqi)

        sheet = excel['订购成功']
        for i in range(31):
            dinggou = sheet.cell(row=18, column=i + 10).value
            day_dinggou.append(dinggou)

        sheet = excel['二次确认']
        for i in range(31):
            erciqueren = sheet.cell(row=18, column=i + 4).value
            day_erciqueren.append(erciqueren)

        sheet = excel['佣金结算表']
        duoribaom = 0
        bingjilingm = 0
        chaohuim = 0
        yuyinbaom = 0
        yuebaom = 0
        jiasubaom = 0
        zunxiangbaom = 0
        for i in range(31):
            for r in range(15):
                if r < 2:
                    duoribaom += sheet.cell(row = r + 2, column = 7 + i).value
                elif r < 4:
                    bingjilingm += sheet.cell(row = r + 2, column = 7 + i).value
                elif r < 6:
                    chaohuim += sheet.cell(row = r + 2, column = 7 + i).value
                elif r < 9:
                    yuyinbaom += sheet.cell(row = r + 2, column = 7 + i).value
                elif r < 12:
                    yuebaom += sheet.cell(row = r + 2, column = 7 + i).value
                elif r < 13:
                    jiasubaom += sheet.cell(row = r + 2, column = 7 + i).value
                else:
                    zunxiangbaom += sheet.cell(row = r + 2, column = 7 + i).value
            duoribao_money.append(duoribaom)
            bingjiling_money.append(bingjilingm)
            chaohui_money.append(chaohuim)
            yuyinbao_money.append(yuyinbaom)
            yuebao_money.append(yuebaom)
            jiasubao_money.append(jiasubaom)
            zunxiangbao_money.append(zunxiangbaom)

        excel.close()
    if not os.path.exists(hljfilename):
        month -= 1
    sqldb.inserthlj(year, month, day_sale, day_money, total_money, duoribao_sale, shengdaibao_sale, chaohui_sale, yuyinbao_sale,yuebao_sale,
                    jiasubao_sale, zunxiangbao_sale,day_faqi,day_erciqueren, day_dinggou, duoribao_money,bingjiling_money,chaohui_money,yuyinbao_money,
                    yuebao_money,jiasubao_money,zunxiangbao_money)


# 复制数据库
def copydb():
    # 数据库路径
    sourcePath = r'D:\pytest\sichangyunLoadnum\sichangyun.db'
    # 复制到目标路径
    targetPath = r'D:\pytest\sichangyun\sichangyun.db'
    shutil.copy(sourcePath, targetPath)


if __name__ == '__main__':
    #获取年月
    year = datetime.datetime.now().year
    month = datetime.datetime.now().month

    #复制当天日报至指定位置
    copyfile(year, month)
    sqldb.deldb(year, month)
    #读取日报数据
    readnum(year, month)
    # 复制数据库
    copydb()


